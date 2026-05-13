import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Dados atualizados com o número da portaria
data = [
    {
        "Portaria": "424/2026",
        "Nome": "SEVERIANO JOSÉ COSTANDRADE DE AGUIAR",
        "Cargo": "Conselheiro",
        "Matricula": "24.003-2",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2505.62,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "STHEFAN BRAVIN PONCHE",
        "Cargo": "Major-PM-TO",
        "Matricula": "27.030-3",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "GEORGES AIRES NUNES",
        "Cargo": "Subtenente QPPM",
        "Matricula": "24.539-8",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "THALYTA MAYANE CARVALHO FERNANDES BRAZ",
        "Cargo": "Assistente de Gabinete de Conselheiro",
        "Matricula": "27.053-6",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2255.06,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "RAIMUNDO NONATO GOMES MONTURIL NETO",
        "Cargo": "Assessor IV",
        "Matricula": "24.305-4",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "WASHINGTON DE SOUSA MELO JÚNIOR",
        "Cargo": "Assessor II",
        "Matricula": "24.587-6",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "AMARILDO CARVALHO DE ALMEIDA",
        "Cargo": "Assessor I",
        "Matricula": "27.006-3",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "DÊNIA MARIA ALMEIDA DA LUZ SOARES",
        "Cargo": "Diretora-Geral de Controle Externo",
        "Matricula": "23.604-7",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "PEDRO HENRIQUE RIBEIRO CASIMIRO",
        "Cargo": "Assessor I",
        "Matricula": "27.043-2",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "SYLLAS FRANKLIN RODRIGUES GOMES",
        "Cargo": "Assessor II",
        "Matricula": "27.032-6",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "KAMILLA SOUSA DE OLIVEIRA",
        "Cargo": "Assessora de Gabinete da Ouvidoria",
        "Matricula": "27.012-6",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "GILSON JOSÉ PEREIRA DOS SANTOS",
        "Cargo": "Assistente de Gabinete da Ouvidoria",
        "Matricula": "27.010-9",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "WELSON GOMES RIBEIRO",
        "Cargo": "Assistente de Controle Externo",
        "Matricula": "23.614-4",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "DHENIA GERHARDT BERNARDON",
        "Cargo": "Assessora Especial de Comunicação",
        "Matricula": "24.394-3",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2255.06,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "LUIZ HENRIQUE PAULINO MACHADO",
        "Cargo": "Assessor I",
        "Matricula": "27.052-0",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "EDIVAN OLIVEIRA CAVALCANTI",
        "Cargo": "Assessor III",
        "Matricula": "27.006-8",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "WANESSA DA SILVA BOTELHO",
        "Cargo": "Assessora de Imprensa e Relações Públicas",
        "Matricula": "27.052-3",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "ERNANI FERREIRA DE CASTRO",
        "Cargo": "Assistente Operacional da Presidência",
        "Matricula": "24.544-3",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "DOURIVAL DA ROCHA MIRANDA FILHO",
        "Cargo": "Cedido",
        "Matricula": "27.026-1",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "LUIS EDUARDO CREMA SIQUEIRA CAMPOS",
        "Cargo": "Assessor I",
        "Matricula": "27.056-8",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "DAGMAR ALBERTINA GEMELLI",
        "Cargo": "Assessora Especial de Gabinete de Conselheiro",
        "Matricula": "23.763-9",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2255.06,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "ADRIANO BARBOSA DE OLIVEIRA",
        "Cargo": "Assessor Especial de Gabinete de Conselheiro",
        "Matricula": "27.002-7",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "424/2026",
        "Nome": "JOÃO CARLOS NUNES BARBOSA",
        "Cargo": "Assistente de Gabinete de Conselheiro",
        "Matricula": "27.001-4",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2044.98,
        "Processo SEI": "26.001937-2"
    },
    {
        "Portaria": "425/2026",
        "Nome": "ALBERTO SEVILHA",
        "Cargo": "Conselheiro Presidente",
        "Matricula": "23.842-2",
        "Data Publicacao": "11/05/2026",
        "Itinerario": "Palmas/TO, Combinado/TO, Arraias/TO, Palmas/TO",
        "Periodo": "11/05/2026 a 14/05/2026",
        "Qtd Diarias": 3.5,
        "Total Diaria": 2505.62,
        "Processo SEI": "26.001937-2"
    }
]

# Criar DataFrame
df = pd.DataFrame(data)

# Criar Workbook do Excel
wb = Workbook()
ws = wb.active
ws.title = "Diárias TCE-TO"

# Definir Estilos
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
body_font = Font(name="Arial", size=10)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Cabeçalhos
headers = ["Portaria", "Nome", "Cargo", "Matrícula", "Data Publicação", "Itinerário", "Período", "Qtd Diárias", "Total Diária", "Processo SEI"]
ws.append(headers)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Adicionar Dados
for index, row in df.iterrows():
    row_data = [
        row["Portaria"], row["Nome"], row["Cargo"], row["Matricula"], row["Data Publicacao"],
        row["Itinerario"], row["Periodo"], row["Qtd Diarias"], row["Total Diaria"],
        row["Processo SEI"]
    ]
    ws.append(row_data)
    
    # Formatação de Moeda e Números
    row_num = index + 2
    ws.cell(row=row_num, column=9).number_format = '"R$ "#,##0.00'
    ws.cell(row=row_num, column=8).number_format = '0.0'
    
    # Estilo do Corpo
    for cell in ws[row_num]:
        cell.font = body_font
        cell.border = border

# Ajustar largura das colunas
column_widths = [15, 45, 35, 12, 18, 50, 25, 12, 15, 18]
for i, width in enumerate(column_widths):
    ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width

# Salvar Arquivo
output_file = "diarias/diario_oficial/diarias_tce_to.xlsx"
wb.save(output_file)
print(f"Planilha atualizada com sucesso em: {output_file}")
