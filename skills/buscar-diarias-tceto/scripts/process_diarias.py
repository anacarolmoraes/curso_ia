import os
import re
import shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Configurações de caminhos
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
DIARIO_DIR = os.path.join(BASE_DIR, "diarias", "diario_oficial")
PROCESSED_DIR = os.path.join(DIARIO_DIR, "processados")
EXCEL_PATH = os.path.join(DIARIO_DIR, "diarias_tce_to.xlsx")
LOG_PATH = os.path.join(DIARIO_DIR, "processados.md")

def get_processed_list():
    """Lê a lista de diários já processados do log markdown."""
    if not os.path.exists(LOG_PATH):
        return []
    with open(LOG_PATH, 'r', encoding='utf-8') as f:
        content = f.read()
    return re.findall(r"- (\d+)", content)

def update_processed_list(bulletin_number):
    """Adiciona um novo número ao log de processados."""
    if not os.path.exists(LOG_PATH):
        with open(LOG_PATH, 'w', encoding='utf-8') as f:
            f.write("# Diários Processados - TCE-TO\n\n")
    
    processed = get_processed_list()
    if str(bulletin_number) not in processed:
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(f"- {bulletin_number}\n")

def check_excel_for_bulletin(bulletin_number):
    """Verifica se o boletim já possui dados na planilha Excel."""
    if not os.path.exists(EXCEL_PATH):
        return False
    try:
        df = pd.read_excel(EXCEL_PATH)
        if "Boletim" in df.columns:
            return str(bulletin_number) in df["Boletim"].astype(str).values
    except:
        pass
    return False

def setup_excel():
    """Cria o arquivo Excel com cabeçalhos se não existir."""
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Diárias TCE-TO"
        headers = ["Boletim", "Portaria", "Nome", "Cargo", "Matrícula", "Data Publicação", "Itinerário", "Período", "Qtd Diárias", "Total Diária", "Processo SEI"]
        ws.append(headers)
        
        # Estilo cabeçalho
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(EXCEL_PATH)

def format_excel():
    """Aplica formatação final na planilha."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    body_font = Font(name="Arial", size=10)
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            # Ajuste de índices por conta da nova coluna Boletim (agora Diárias é 9 e Total é 10)
            if cell.column == 9:
                cell.number_format = '0.0'
            elif cell.column == 10:
                cell.number_format = '"R$ "#,##0.00'
                
    # Ajustar larguras
    column_widths = [12, 15, 45, 35, 12, 18, 50, 25, 12, 15, 18]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width
        
    wb.save(EXCEL_PATH)

def extract_data_from_md(file_path):
    """Extrai portarias de viagem do arquivo markdown."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    extracted_records = []
    
    # Extrair número do boletim do nome do arquivo ou conteúdo
    # Ex nome arquivo: BO_3948_11-05-2026.md
    file_name = os.path.basename(file_path)
    boletim_match = re.search(r"BO_(\d+)", file_name)
    boletim_num = boletim_match.group(1) if boletim_match else "N/A"

    # Extrair data de publicação do boletim
    pub_date_match = re.search(r"Publicado em (\d{2}/\d{2}/\d{4})", content)
    pub_date = pub_date_match.group(1) if pub_date_match else "Desconhecida"
    
    # Dividir por portarias
    portarias = re.split(r"# PORTARIA Nº", content)
    
    for port_content in portarias[1:]:
        port_num_match = re.search(r"(\d+/\d{4})", port_content)
        if not port_num_match: continue
        port_num = port_num_match.group(1)
        
        # Filtrar estritamente: Deve ser da presidência e autorizar viagem ou diárias
        port_lower = port_content.lower()
        is_from_president = "presidente" in port_lower or "presidência" in port_lower
        is_authorizing = any(word in port_lower for word in ["autorizar", "conceder", "atribuir", "designar"])
        is_travel_or_per_diem = any(term in port_lower for term in ["viagem", "diária", "deslocamento", "itinerário"])
        
        if is_from_president and is_authorizing and is_travel_or_per_diem:
            # Extrair Processo SEI
            sei_match = re.search(r"Processo SEI nº\s*([\d\.\-]+)", port_content)
            sei_proc = sei_match.group(1) if sei_match else "Não informado"
            
            # Extrair Itinerário e Período
            itinerario_match = re.search(r"\*\*Itinerário:\s*\*\*(.*?)(?=\*\*|$)", port_content, re.DOTALL | re.IGNORECASE)
            periodo_match = re.search(r"\*\*Período:\s*\*\*(.*?)(?=\*\*|$)", port_content, re.DOTALL | re.IGNORECASE)
            
            itinerario = itinerario_match.group(1).strip() if itinerario_match else ""
            periodo_full = periodo_match.group(1).strip() if periodo_match else ""
            
            # Extrair Diárias (ex: 3,5 diárias)
            qtd_match = re.search(r"([\d,\.]+)\s*\(.*?\)\s*diárias", port_content)
            qtd_diarias = float(qtd_match.group(1).replace(',', '.')) if qtd_match else 0.0
            
            # Isolar a parte dos servidores (geralmente após o Período) para evitar preâmbulos
            part_after_periodo = port_content
            if "**Período:**" in port_content:
                part_after_periodo = port_content.split("**Período:**")[-1]
            elif "Período:" in port_content:
                part_after_periodo = port_content.split("Período:")[-1]
            elif "**Itinerário:**" in port_content:
                part_after_periodo = port_content.split("**Itinerário:**")[-1]

            # Regex aprimorada para focar apenas nos nomes e limitar o tamanho do cargo
            regex_pessoas = r"([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ\s]+),\s*(.{1,100}?),\s*mat\.?\s*(?:nº|n\.)?\s*([\d\.\-]+).*?R\$\s*([\d\.,]+)"
            people_matches = re.finditer(regex_pessoas, part_after_periodo, re.DOTALL)
            
            for m in people_matches:
                nome = m.group(1).strip()
                cargo = m.group(2).strip()
                matricula = m.group(3).strip()
                valor_diaria = float(m.group(4).replace('.', '').replace(',', '.'))
                
                # Tratar caso especial onde o cargo vem como asteriscos (ex: Presidente)
                if cargo in ['** **', '**', ''] or "Presidente" in port_content:
                    # Se for o presidente, não tem cargo escrito claramente
                    if "SEVILHA" in nome or "PRESIDENTE" in port_content:
                         cargo = "Conselheiro Presidente"
                         
                extracted_records.append({
                    "Boletim": boletim_num,
                    "Portaria": port_num,
                    "Nome": nome,
                    "Cargo": cargo,
                    "Matrícula": matricula,
                    "Data Publicação": pub_date,
                    "Itinerário": itinerario,
                    "Período": periodo_full,
                    "Qtd Diárias": qtd_diarias,
                    "Total Diária": valor_diaria,
                    "Processo SEI": sei_proc
                })

    return extracted_records

def main():
    if not os.path.exists(PROCESSED_DIR):
        os.makedirs(PROCESSED_DIR)
    
    setup_excel()
    
    files = [f for f in os.listdir(DIARIO_DIR) if f.endswith(".md") and f != "processados.md"]
    
    if not files:
        print("Nenhum arquivo .md novo encontrado para processar.")
        return

    all_new_data = []
    processed_bulletins = []

    for file_name in files:
        # Extrair número do boletim para verificação
        boletim_match = re.search(r"BO_(\d+)", file_name)
        bulletin_num = boletim_match.group(1) if boletim_match else None
        
        if bulletin_num and check_excel_for_bulletin(bulletin_num):
            print(f"[!] AVISO: Dados do Boletim {bulletin_num} já existem no Excel. Movendo para 'processados' sem extrair.")
            shutil.move(os.path.join(DIARIO_DIR, file_name), os.path.join(PROCESSED_DIR, file_name))
            update_processed_list(bulletin_num)
            continue

        file_path = os.path.join(DIARIO_DIR, file_name)
        print(f"Processando: {file_name}")
        
        data = extract_data_from_md(file_path)
        all_new_data.extend(data)
        
        if bulletin_num:
            processed_bulletins.append(bulletin_num)
        
        # Mover para processados
        shutil.move(file_path, os.path.join(PROCESSED_DIR, file_name))
        print(f"Arquivo movido para pasta 'processados'.")

    if all_new_data:
        df_new = pd.DataFrame(all_new_data)
        
        try:
            df_existing = pd.read_excel(EXCEL_PATH)
            df_final = pd.concat([df_existing, df_new]).drop_duplicates(subset=["Portaria", "Nome"], keep='last')
        except:
            df_final = df_new

        with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name="Diárias TCE-TO")
        
        format_excel()
        
        # Atualizar log para cada boletim processado com sucesso
        for b in processed_bulletins:
            update_processed_list(b)
            
        print(f"Processamento concluído. {len(all_new_data)} registros novos/atualizados.")
    else:
        print("Nenhum registro novo de diária foi adicionado.")

if __name__ == "__main__":
    main()
